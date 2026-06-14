import csv
import re
from pathlib import Path
from openpyxl import load_workbook


RAW_FILE = Path("data/elt_raw.xlsx")
OUTPUT_FILE = Path("data/elt_lienholders_clean.csv")
REVIEW_FILE = Path("data/elt_lienholders_duplicates_review.csv")


SOURCE_NAME = "Florida ELT / Lienholder Source List"


COLUMN_MAP = {
    "FULL NAME": "display_name",
    "M STREETADDRESS": "mailing_address",
    "M CITY": "city",
    "M STATECODE": "state",
    "M ZIPCODE": "zip",
    "FEIDNUMBER": "feid_number",
    "FEIDSUFFIX": "feid_suffix",
    "CUSTOMERNUMBER": "elt_customer_number",
}


FINAL_COLUMNS = [
    "display_name",
    "search_name",
    "mailing_address",
    "city",
    "state",
    "zip",
    "feid_number",
    "feid_suffix",
    "elt_customer_number",
    "phone",
    "email",
    "payoff_website",
    "lien_release_notes",
    "elt_notes",
    "search_aliases",
    "last_verified_date",
    "internal_notes",
    "source",
]


def clean_text(value):
    """Basic safe text cleanup."""
    if value is None:
        return ""

    value = str(value).strip()

    # Convert Excel-looking floats like 12345.0 to 12345
    if re.fullmatch(r"\d+\.0", value):
        value = value[:-2]

    # Normalize whitespace
    value = re.sub(r"\s+", " ", value)

    return value


def clean_name(value):
    """Clean lienholder display name without over-changing legal names."""
    value = clean_text(value)
    value = value.upper()

    # Normalize punctuation spacing
    value = re.sub(r"\s*,\s*", ", ", value)
    value = re.sub(r"\s*\.\s*", ". ", value)
    value = re.sub(r"\s+", " ", value).strip()

    # Remove trailing space after periods if it creates odd endings
    value = value.replace(" ,", ",").strip()

    return value


def make_search_name(display_name):
    """Create forgiving search version of the lienholder name."""
    name = display_name.upper()

    # Replace punctuation with spaces
    name = re.sub(r"[^A-Z0-9]+", " ", name)

    # Remove common noise words only from search field, not display name
    noise_words = {
        "THE",
        "INC",
        "INCORPORATED",
        "LLC",
        "L L C",
        "CORP",
        "CORPORATION",
        "CO",
        "COMPANY",
        "LTD",
        "LIMITED",
    }

    parts = [part for part in name.split() if part not in noise_words]
    return " ".join(parts).strip()


def clean_identifier(value):
    """Keep IDs as text."""
    value = clean_text(value)

    # Remove spaces inside ID-like values
    value = re.sub(r"\s+", "", value)

    # Convert Excel-looking float IDs to integer string
    if re.fullmatch(r"\d+\.0", value):
        value = value[:-2]

    return value


def main():
    if not RAW_FILE.exists():
        raise FileNotFoundError(
            f"Could not find {RAW_FILE}. Make sure elt_raw.xlsx is inside the data folder."
        )

    print(f"Reading: {RAW_FILE}")

    wb = load_workbook(RAW_FILE, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)

    raw_headers = next(rows)
    headers = [clean_text(h).upper() for h in raw_headers]

    missing = [col for col in COLUMN_MAP if col not in headers]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}")

    col_indexes = {
        COLUMN_MAP[col]: headers.index(col)
        for col in COLUMN_MAP
    }

    cleaned_rows = []
    seen_keys = set()
    duplicate_review = {}

    total_rows = 0
    kept_rows = 0
    duplicate_rows = 0

    for row in rows:
        total_rows += 1

        display_name = clean_name(row[col_indexes["display_name"]])
        if not display_name:
            continue

        mailing_address = clean_text(row[col_indexes["mailing_address"]]).upper()
        city = clean_text(row[col_indexes["city"]]).upper()
        state = clean_text(row[col_indexes["state"]]).upper()
        zip_code = clean_identifier(row[col_indexes["zip"]])
        feid_number = clean_identifier(row[col_indexes["feid_number"]])
        feid_suffix = clean_identifier(row[col_indexes["feid_suffix"]])
        elt_customer_number = clean_identifier(row[col_indexes["elt_customer_number"]])

        search_name = make_search_name(display_name)

        dedupe_key = (
            search_name,
            mailing_address,
            city,
            state,
            zip_code,
            feid_number,
            feid_suffix,
            elt_customer_number,
        )

        duplicate_review.setdefault(search_name, 0)
        duplicate_review[search_name] += 1

        if dedupe_key in seen_keys:
            duplicate_rows += 1
            continue

        seen_keys.add(dedupe_key)

        cleaned_rows.append({
            "display_name": display_name,
            "search_name": search_name,
            "mailing_address": mailing_address,
            "city": city,
            "state": state,
            "zip": zip_code,
            "feid_number": feid_number,
            "feid_suffix": feid_suffix,
            "elt_customer_number": elt_customer_number,
            "phone": "",
            "email": "",
            "payoff_website": "",
            "lien_release_notes": "",
            "elt_notes": "",
            "search_aliases": "",
            "last_verified_date": "",
            "internal_notes": "",
            "source": SOURCE_NAME,
        })

        kept_rows += 1

    cleaned_rows.sort(key=lambda x: (x["display_name"], x["state"], x["city"]))

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    with OUTPUT_FILE.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FINAL_COLUMNS)
        writer.writeheader()
        writer.writerows(cleaned_rows)

    review_rows = [
        {"search_name": name, "record_count": count}
        for name, count in duplicate_review.items()
        if count > 1
    ]
    review_rows.sort(key=lambda x: x["record_count"], reverse=True)

    with REVIEW_FILE.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["search_name", "record_count"])
        writer.writeheader()
        writer.writerows(review_rows)

    print("Cleanup complete.")
    print(f"Raw rows read: {total_rows}")
    print(f"Clean records kept: {len(cleaned_rows)}")
    print(f"Duplicate exact rows removed: {duplicate_rows}")
    print(f"Clean CSV created: {OUTPUT_FILE}")
    print(f"Duplicate review CSV created: {REVIEW_FILE}")


if __name__ == "__main__":
    main()